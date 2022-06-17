import openpyxl
from configparser import ConfigParser
import priceCrawler
import crawler.akshare_api as ak
from mathhelp.irr import xirr
import xlrd

date_now = 0

def get_headers(sheet):
    col_num = sheet.max_column
    headers = []
    for i in range(1, col_num + 1):
        headers.append(sheet.cell(row=1, column=i).value)
    return headers

def latest_price(code, exchange):
    global date_now
    # 未上市
    if code in {"688297"}:
        return 32.35

    if exchange in {"上海", "深圳", "香港", "新三板", "新股"}:
        return ak.get_stock_price(code)
    elif exchange in {"可转债"}:
        return ak.get_bond_price(code)
    elif exchange in {"场内基金"}:
        return ak.get_etf_fund_price(code)
    elif exchange in {"场外基金"}:
        nav, date_now = priceCrawler.get_fund_price(code)
        return nav
    elif exchange in {"宽基"}:
        return priceCrawler.get_stock_price("SH." + code, "")
    elif exchange in {"现金"}:
        return 1
    else:
        return -1


def update_peterlynch_price(excelpath):
    wb = openpyxl.load_workbook(excelpath, data_only=False)

    # 更新汇率
    config_sheet = wb.get_sheet_by_name("参数")
    config_sheet.cell(row=1, column=2).value = priceCrawler.get_hk_rate()

    # 保护数据透视表格式
    pivot_sheet = wb.get_sheet_by_name("总计")
    pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
    pivot.cache.refreshOnLoad = True  # 刷新加载

    # 更新股价
    data_sheet = wb.get_sheet_by_name("整体估值")
    headers = get_headers(data_sheet)
    code_col = headers.index('代码') + 1
    price_col = headers.index('股价') + 1
    category_col = headers.index('行业分类') + 1

    row_num = data_sheet.max_row
    for i in range(2, row_num + 1):
        print(data_sheet.cell(i, code_col).value)
        if data_sheet.cell(i, code_col).value is None:
            pass
        elif data_sheet.cell(i, category_col).value == "宽基":
            nav = latest_price(data_sheet.cell(i, code_col).value, "宽基")
            data_sheet.cell(i, price_col).value = nav
        else:
            nav = latest_price(data_sheet.cell(i, code_col).value, "上海")
            data_sheet.cell(i, price_col).value = nav

    # 更新报表好看一列
    data_sheet2 = wb.get_sheet_by_name("报表好看")
    headers2 = get_headers(data_sheet2)
    code_col = headers2.index('代码') + 1
    price_col = headers2.index('现价') + 1

    row_num = data_sheet2.max_row
    for i in range(2, row_num + 1):
        nav = latest_price(str(data_sheet2.cell(i, code_col).value), "上海")
        data_sheet2.cell(i, price_col).value = nav

    wb.save(excelpath)


def update_position_price(excelpath):
    wb = openpyxl.load_workbook(excelpath, data_only=False)

    # 更新汇率
    config_sheet = wb.get_sheet_by_name("参数")
    config_sheet.cell(row=1, column=2).value = priceCrawler.get_hk_rate()

    # 保护数据透视表格式
    pivot_sheet = wb.get_sheet_by_name("总计")
    pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
    pivot.cache.refreshOnLoad = True  # 刷新加载

    pivot_sheet2 = wb.get_sheet_by_name("股票总计")
    pivot2 = pivot_sheet2._pivots[0]  # 任何一个都可以共享同一个缓存
    pivot2.cache.refreshOnLoad = True  # 刷新加载

    # 更新股价
    data_sheet = wb.get_sheet_by_name("实时持仓")
    headers = get_headers(data_sheet)
    code_col = headers.index('code') + 1
    exchange_col = headers.index('交易板块') + 1
    price_col = headers.index('股价') + 1

    row_num = data_sheet.max_row
    for i in range(2, row_num + 1):
        nav = latest_price(data_sheet.cell(i, code_col).value, data_sheet.cell(i, exchange_col).value)
        data_sheet.cell(i, price_col).value = nav
    wb.save(excelpath)


def update_stock_price(excelpath):
    wb = openpyxl.load_workbook(excelpath, data_only=False)

    # 更新汇率
    config_sheet = wb.get_sheet_by_name("参数")
    config_sheet.cell(row=1, column=2).value = priceCrawler.get_hk_rate()

    # 更新股价
    data_sheet = wb.get_sheet_by_name("当前持仓")
    headers = get_headers(data_sheet)
    code_col = headers.index('股票代码') + 1
    price_col = headers.index('现价') + 1
    inc_col = [headers.index('一日涨幅') + 1, headers.index('五日涨幅') + 1, headers.index('近一月') + 1,
                   headers.index('近半年') + 1, headers.index('近一年') + 1]

    row_num = data_sheet.max_row
    for i in range(2, row_num + 1):
        nav = latest_price(data_sheet.cell(i, code_col).value, "上海")
        data_sheet.cell(i, price_col).value = nav

        inc_list = priceCrawler.get_sina_increase(data_sheet.cell(i, code_col).value)
        if inc_list[0] != -200:
            for j in range(len(inc_col)):
                if inc_list[j] == -100:
                    data_sheet.cell(i, inc_col[j]).value = "--"
                else:
                    data_sheet.cell(i, inc_col[j]).value = inc_list[j]

    # 更新债价
    data_sheet2 = wb.get_sheet_by_name("垃圾转债")
    headers2 = get_headers(data_sheet2)
    code_col = headers2.index('股票代码') + 1
    price_col = headers2.index('现价') + 1
    rate_col = headers2.index('到期年化') + 1
    duedate_col = headers2.index('到期日期') + 1

    row_num = data_sheet2.max_row
    for i in range(2, row_num + 1):
        nav = latest_price(str(data_sheet2.cell(i, code_col).value), "可转债")
        data_sheet2.cell(i, price_col).value = nav

        mytas = []
        mytas.append((date_now, - nav))
        mytas.append((data_sheet2.cell(i, duedate_col).value, 100))
        data_sheet2.cell(i, rate_col).value = xirr(mytas)
    wb.save(excelpath)


def update_stock_basic_info(excelpath, a_shareinfo, h_shareinfo):
    wb = openpyxl.load_workbook(excelpath, data_only=False)
    data_sheet = wb.get_sheet_by_name("当前持仓")
    headers = get_headers(data_sheet)
    print(headers)
    code_col = headers.index('股票代码') + 1
    industry_col = headers.index('一级行业') + 1
    equity_col = headers.index('总股本') + 1
    foreign_col = headers.index('外资比') + 1
    report2yuan_col = headers.index('财报货币-人民币') + 1
    price2yuan_col = headers.index('股价货币-人民币') + 1
    netprofit_col = headers.index('预测净利润(元)2023') + 1
    grossmargin_col = headers.index('销售毛利率2020三季度') + 1
    goodwillrate_col = headers.index('商誉净利润比2020三季度') + 1
    roic_col = headers.index('投入资本回报率(%)') + 1
    inc_col = headers.index('一日涨幅') + 1

    # 记录股票位置
    code2row_dict = {}
    row_num = data_sheet.max_row
    for i in range(2, row_num + 1):
        code2row_dict[data_sheet.cell(i, code_col).value] = i

    # 更新A股基础信息
    wb_a = openpyxl.load_workbook(a_shareinfo, data_only=False)
    data_sheet_a = wb_a.get_sheet_by_name("选股结果")
    headers_a = get_headers(data_sheet_a)
    print(headers_a)
    code_col_a = headers_a.index('股票代码') + 1
    industry_col_a = headers_a.index('所属申万行业') + 1
    netprofit_col_a = headers_a.index('预测净利润(元)') + 1
    grossmargin_col_a = headers_a.index('销售毛利率(%)') + 1
    goodwillrate_col_a = headers_a.index('[1] / [2]') + 1 #headers_a.index('{(}商誉{/}净利润{)}') + 1
    roic_col_a = -1
    equity_col_a = -1
    foreign_col_a = -1
    for i,x in enumerate(headers_a):
        if x is None:
            continue
        elif x.startswith('总股本'):
            equity_col_a = i + 1
        elif x.startswith('陆股通持股占总股本比'):
            foreign_col_a = i + 1
        elif x.startswith('投入资本回报率(%)'):
            roic_col_a = i + 1

    row_num_a = data_sheet_a.max_row
    for i in range(3, row_num_a + 1):
        if data_sheet_a.cell(i, code_col_a).value[:6] in code2row_dict.keys():
            print("update " + data_sheet_a.cell(i, code_col_a).value[:6])
            i_origin = code2row_dict[data_sheet_a.cell(i, code_col_a).value[:6]]
            industry_name = data_sheet_a.cell(i, industry_col_a).value.split('--')
            print(i_origin, industry_name, data_sheet_a.cell(i, equity_col_a).value,
                  data_sheet_a.cell(i, foreign_col_a).value, data_sheet_a.cell(i, netprofit_col_a).value,
                  data_sheet_a.cell(i, netprofit_col_a + 1).value, data_sheet_a.cell(i, netprofit_col_a + 2).value,
                  data_sheet_a.cell(i, netprofit_col_a + 3).value,
                  data_sheet_a.cell(i, grossmargin_col_a).value, data_sheet_a.cell(i, grossmargin_col_a + 1).value,
                  data_sheet_a.cell(i, goodwillrate_col_a).value, data_sheet_a.cell(i, goodwillrate_col_a + 1).value,
                  data_sheet_a.cell(i, roic_col_a).value)
            data_sheet.cell(i_origin, industry_col).value = industry_name[0]
            data_sheet.cell(i_origin, industry_col + 1).value = industry_name[1]
            data_sheet.cell(i_origin, industry_col + 2).value = industry_name[2]
            data_sheet.cell(i_origin, equity_col).value = data_sheet_a.cell(i, equity_col_a).value
            data_sheet.cell(i_origin, foreign_col).value = data_sheet_a.cell(i, foreign_col_a).value
            data_sheet.cell(i_origin, report2yuan_col).value = 1
            data_sheet.cell(i_origin, price2yuan_col).value = 1
            data_sheet.cell(i_origin, netprofit_col).value = data_sheet_a.cell(i, netprofit_col_a).value
            data_sheet.cell(i_origin, netprofit_col + 1).value = data_sheet_a.cell(i, netprofit_col_a + 1).value
            data_sheet.cell(i_origin, netprofit_col + 2).value = data_sheet_a.cell(i, netprofit_col_a + 2).value
            data_sheet.cell(i_origin, netprofit_col + 3).value = data_sheet_a.cell(i, netprofit_col_a + 3).value
            data_sheet.cell(i_origin, grossmargin_col).value = data_sheet_a.cell(i, grossmargin_col_a).value
            data_sheet.cell(i_origin, grossmargin_col + 1).value = data_sheet_a.cell(i, grossmargin_col_a + 1).value
            data_sheet.cell(i_origin, goodwillrate_col).value = data_sheet_a.cell(i, goodwillrate_col_a).value
            data_sheet.cell(i_origin, goodwillrate_col + 1).value = data_sheet_a.cell(i, goodwillrate_col_a + 1).value
            data_sheet.cell(i_origin, roic_col).value = data_sheet_a.cell(i, roic_col_a).value
    wb_a.close()

    # 更新港股基础信息
    wb_b = openpyxl.load_workbook(h_shareinfo, data_only=False)
    data_sheet_b = wb_b.get_sheet_by_name("选股结果")
    headers_b = get_headers(data_sheet_b)
    #print(headers_b)
    code_col_b = headers_b.index('股票代码') + 1
    industry_col_b = headers_b.index('所属恒生行业') + 1
    netprofit_col_b = headers_b.index('预测净利润平均值(港元)') + 1
    grossmargin_col_b = headers_b.index('销售毛利率(%)') + 1
    goodwillrate_col_b = headers_b.index('商誉(港元)') + 1
    roic_col_b = -1
    inc_col_b = -1
    equity_col_b = -1
    for i, x in enumerate(headers_b):
        if x is None:
            continue
        elif x.startswith('涨跌幅'):
            inc_col_b = i + 1
        elif x.startswith('总股本'):
            equity_col_b = i + 1
        elif x.startswith('投入资本回报率(%)'):
            roic_col_b = i + 1

    row_num_b = data_sheet_b.max_row
    for i in range(3, row_num_b + 1):
        if str("hk0" + data_sheet_b.cell(i, code_col_b).value[:4]) in code2row_dict.keys():
            print("update " + data_sheet_b.cell(i, code_col_b).value[:4])
            i_origin = code2row_dict["hk0" + data_sheet_b.cell(i, code_col_b).value[:4]]
            print(i_origin, data_sheet_b.cell(i, industry_col_b).value,
                  data_sheet_b.cell(i, industry_col_b + 1).value, data_sheet_b.cell(i, industry_col_b + 2).value,
                  data_sheet_b.cell(i, equity_col_b).value,
                  data_sheet_b.cell(i, netprofit_col_b).value, data_sheet_b.cell(i, netprofit_col_b + 1).value,
                  data_sheet_b.cell(i, netprofit_col_b + 2).value, data_sheet_b.cell(i, netprofit_col_b + 3).value,
                  data_sheet_b.cell(i, grossmargin_col_b).value, data_sheet_b.cell(i, grossmargin_col_b + 1).value,
                  data_sheet_b.cell(i, goodwillrate_col_b).value, data_sheet_b.cell(i, goodwillrate_col_b + 1).value,
                  data_sheet_b.cell(i, inc_col_b).value,
                  data_sheet_b.cell(i, inc_col_b + 1).value, data_sheet_b.cell(i, inc_col_b + 2).value,
                  data_sheet_b.cell(i, inc_col_b + 3).value, data_sheet_b.cell(i, inc_col_b + 4).value,
                  data_sheet_b.cell(i, roic_col_b).value)
            data_sheet.cell(i_origin, industry_col).value = data_sheet_b.cell(i, industry_col_b).value
            data_sheet.cell(i_origin, industry_col + 1).value = data_sheet_b.cell(i, industry_col_b + 1).value
            data_sheet.cell(i_origin, industry_col + 2).value = data_sheet_b.cell(i, industry_col_b + 2).value
            data_sheet.cell(i_origin, equity_col).value = data_sheet_b.cell(i, equity_col_b).value
            data_sheet.cell(i_origin, netprofit_col).value = data_sheet_b.cell(i, netprofit_col_b).value
            data_sheet.cell(i_origin, netprofit_col + 1).value = data_sheet_b.cell(i, netprofit_col_b + 1).value
            data_sheet.cell(i_origin, netprofit_col + 2).value = data_sheet_b.cell(i, netprofit_col_b + 2).value
            data_sheet.cell(i_origin, netprofit_col + 3).value = data_sheet_b.cell(i, netprofit_col_b + 3).value
            data_sheet.cell(i_origin, grossmargin_col).value = data_sheet_b.cell(i, grossmargin_col_b).value
            data_sheet.cell(i_origin, grossmargin_col + 1).value = data_sheet_b.cell(i, grossmargin_col_b + 1).value
            data_sheet.cell(i_origin, goodwillrate_col).value = data_sheet_b.cell(i, goodwillrate_col_b).value
            data_sheet.cell(i_origin, goodwillrate_col + 1).value = data_sheet_b.cell(i, goodwillrate_col_b + 1).value
            data_sheet.cell(i_origin, roic_col).value = data_sheet_b.cell(i, roic_col_b).value
            data_sheet.cell(i_origin, inc_col).value = data_sheet_b.cell(i, inc_col_b).value
            data_sheet.cell(i_origin, inc_col + 1).value = data_sheet_b.cell(i, inc_col_b + 1).value
            data_sheet.cell(i_origin, inc_col + 2).value = data_sheet_b.cell(i, inc_col_b + 2).value
            data_sheet.cell(i_origin, inc_col + 3).value = data_sheet_b.cell(i, inc_col_b + 3).value
            data_sheet.cell(i_origin, inc_col + 4).value = data_sheet_b.cell(i, inc_col_b + 4).value
    wb_b.close()

    wb.save(excelpath)

def update_hkrate(excelpath):
    wb = openpyxl.load_workbook(excelpath, data_only=False)
    sheet1 = wb.get_sheet_by_name("参数")  # 通过索引获取表格，一个文件里可能有多个sheet
    sheet1.cell(row=1, column=2).value = priceCrawler.get_hk_rate()

    # 保护数据透视表格式
    pivot_sheet = wb.get_sheet_by_name("总计")
    pivot = pivot_sheet._pivots[0]  # 任何一个都可以共享同一个缓存
    pivot.cache.refreshOnLoad = True  # 刷新加载
    wb.save(excelpath)

if __name__ == '__main__':
    #ak.get_all()

    cfg = ConfigParser()
    cfg.read('config.ini', encoding='UTF-8')
    filepath = cfg.get('file', 'peterlynch_path')
    update_peterlynch_price(filepath)

    filepath = cfg.get('file', 'position_path')
    update_position_price(filepath)

    filepath = cfg.get('file', 'stock_path')
    update_stock_price(filepath)

    # filepath = cfg.get('file', 'stock_path')
    # a_share_info_path1 = cfg.get('file', 'a_share_info_path1')
    # h_share_info_path1 = cfg.get('file', 'h_share_info_path1')
    # update_stock_basic_info(filepath, a_share_info_path1, h_share_info_path1)
    #
    # a_share_info_path2 = cfg.get('file', 'a_share_info_path2')
    # h_share_info_path2 = cfg.get('file', 'h_share_info_path2')
    # update_stock_basic_info(filepath, a_share_info_path2, h_share_info_path2)
    #
    # a_share_info_path3 = cfg.get('file', 'a_share_info_path3')
    # update_stock_basic_info(filepath, a_share_info_path3, h_share_info_path2)
    #
    # a_share_info_path4 = cfg.get('file', 'a_share_info_path4')
    # update_stock_basic_info(filepath, a_share_info_path4, h_share_info_path2)
